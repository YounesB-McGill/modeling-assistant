#!/usr/bin/env python3
"""
Module to help evaluate the Mistake Detection System by parsing the spreadsheets its tests generate and combining the
results in a common spreadsheet.

Author: Younes Boubekeur
"""
# pylint: disable=consider-using-enumerate, wrong-import-position

import sys
import os

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

# this line is needed to be able to call this script directly
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from classdiagram import (AssociationEnd, ClassDiagram, Classifier, CDArray, CDBoolean, CDDouble, CDFloat, CDInt,
                          CDLong, CDString, CDAny, CDByte, CDChar)

from envvars import env_vars
from fileserdes import load_cdm


SPREADSHEET_LOCATIONS: str = env_vars["mds-evaluation-spreadsheet-location"]
COMMON_SPREADSHEET_PATH = f"{SPREADSHEET_LOCATIONS}/common-spreadsheet.xlsx"
INST_CDM_PATH = "mistakedetection/realModels/instructorSolution/smarthome"

INSTRUCTOR_SOLUTION_IDS = (201, 202, 203, 204)
EVALUATED_STUDENT_SOLUTION_IDS = [14, 36, 39, 48, 58, 61, 69, 71, 77, 80, 97, 100]

TN_TP_FP_FN_R_P_F1_F2_M = tuple[int, int, int, int, float, float, float, float, int]  # pylint: disable=invalid-name
MDS_RESULT_COLUMN_OFFSET = 4  # Column E using zero-indexing
INSTRUCTOR_SOLUTION_ID_OFFSET = INSTRUCTOR_SOLUTION_IDS[0] - 1

OVERWRITE_COMMON_SPREADSHEET = False  # default is False to avoid overwriting the file by mistake


def produce_common_spreadsheet():
    "Produce the common results spreadsheet from the individual ones used for the evaluation."
    results = [[(
        "True Negative",
        "True Positive",
        "False Positive",
        "False Negative",
        "Recall (TP / (TP + FN))",
        "Precision (TP / (TP + FP))",
        "F1 (2PR / (P + R))",
        "F2 (5PR / (4P + R)",
        "Number of MDS Mistakes",
    )] for _ in range(len(INSTRUCTOR_SOLUTION_IDS) + 1)]
    for i, instructor_solution_id in enumerate(INSTRUCTOR_SOLUTION_IDS):
        for student_solution_id in EVALUATED_STUDENT_SOLUTION_IDS:
            results[i].append(get_spreadsheet_results(student_solution_id, instructor_solution_id))
        averages = [0 for _ in range(len(results[i][0]))]
        for t in results[i][1:]:
            for j in range(len(averages)):
                averages[j] += t[j] or 0
        for j in range(len(averages)):
            averages[j] /= (len(results[i]) - 1)
        results[i].append(tuple(averages))

    # determine best instructor solution based on lowest number of MDS mistakes
    best_inst_sols = []
    for col in range(1, len(results[0])):
        num_mistakes = []
        for inst_sol_group in range(len(results) - 1):
            num_mistakes.append(results[inst_sol_group][col][-1] or 0)
        best_inst_sol_idx = num_mistakes.index(min(num_mistakes))
        best_inst_sols.append(best_inst_sol_idx + INSTRUCTOR_SOLUTION_ID_OFFSET + 1)
        results[-1].append(results[best_inst_sol_idx][col])

    print_results(results, best_inst_sols)
    if OVERWRITE_COMMON_SPREADSHEET:
        save_to_spreadsheet(results, best_inst_sols, COMMON_SPREADSHEET_PATH)


def save_to_spreadsheet(results: list[list[tuple]], best_inst_sols, path):
    "Saves the given results to the given file path."
    wb = Workbook()
    ws: Worksheet = wb.active
    for i in range(len(INSTRUCTOR_SOLUTION_IDS) + 1):
        if i < len(INSTRUCTOR_SOLUTION_IDS):
            ws.append((f"Instructor solution {INSTRUCTOR_SOLUTION_IDS[i]}",))
        else:
            ws.append(("Best solution based on lowest number of mistakes",))
        ws.append(["Student solution", *EVALUATED_STUDENT_SOLUTION_IDS, "Averages"])
        rows = []
        for j in range(len(results[i][0])):
            rows.append([results[i][k][j] for k in range(len(results[i]))])
        for row in rows:
            ws.append(row)
        if i == len(INSTRUCTOR_SOLUTION_IDS):
            ws.append(["Selected solution", *best_inst_sols])
        ws.append(("",))
    ws.column_dimensions["A"].width *= 2  # increase width of first column to fit text
    wb.save(path)


def print_results(results: list[list[tuple]], best_inst_sols: list[int]):
    "Print the results to console, useful for debugging."
    for i in range(len(INSTRUCTOR_SOLUTION_IDS) + 1):
        if i < len(INSTRUCTOR_SOLUTION_IDS):
            print(f"Instructor solution {INSTRUCTOR_SOLUTION_IDS[i]}")
        else:
            print("Best solution based on lowest number of mistakes")
        print(" ".join(["Student solution", *map(str, EVALUATED_STUDENT_SOLUTION_IDS), "Averages"]))
        for j in range(len(results[i][0])):
            for k in range(len(results[i])):
                print(f"{results[i][k][j]}", end=" ")
            print()
        if i == len(INSTRUCTOR_SOLUTION_IDS):
            print(" ".join(["Selected solution", *map(str, best_inst_sols)]))
        print()


def print_results_as_latex(results: list[list[tuple]], best_inst_sols: list[int]):
    "Print the results to console in LaTeX format, useful for publications."
    for i in range(len(INSTRUCTOR_SOLUTION_IDS) + 1):
        if i < len(INSTRUCTOR_SOLUTION_IDS):
            print(f"Instructor solution {INSTRUCTOR_SOLUTION_IDS[i]}")
        else:
            print("Best solution based on lowest number of mistakes")
        print(" & ".join(["Student solution", *map(str, EVALUATED_STUDENT_SOLUTION_IDS), "Averages"]))
        for j in range(len(results[i][0])):
            for k in range(len(results[i])):
                v = results[i][k][j]
                if isinstance(v, float):
                    v = round(v, 2)
                print(v, end=" & ")
            print()
        if i == len(INSTRUCTOR_SOLUTION_IDS):
            print(" & ".join(["Selected solution", *map(str, best_inst_sols)]))
        print()


def get_spreadsheet_results(student_solution_id, instructor_solution_id) -> TN_TP_FP_FN_R_P_F1_F2_M:
    """
    Return the 9-tuple with the following elements for the given submission_id:

    - True Negatives
    - True Positives
    - False Positives
    - False Negatives
    - Recall (TP / (TP + FN))
    - Precision (TP / (TP + FP))
    - F1 (2 * Precision * Recall / (Precision + Recall))
    - F2 (5 * Precision * Recall) / (4 * Precision + Recall)
    - Number of MDS Mistakes
    """
    result = []
    ws = load_workbook(f"{SPREADSHEET_LOCATIONS}/{student_solution_id}/"
                       f"GroundTruth_{student_solution_id}_{instructor_solution_id}.xlsx",
                       read_only=True, data_only=True).active
    num_mistakes = 0
    reached_tn = False
    for row in ws.values:
        if not reached_tn and row[0]:
            if row[0].startswith("Total Mistakes"):
                num_mistakes = row[MDS_RESULT_COLUMN_OFFSET]
            if row[0].startswith("True Negative"):
                reached_tn = True
        if reached_tn:
            result.append(row[2] if row[2] is not None else 1)  # 1 is a default value if actual value is unavailable
    result.append(num_mistakes)
    return tuple(result)


def print_inst_sol_num_elements():
    "Print the number of elements for each instructor solution."
    for inst_sol_id in INSTRUCTOR_SOLUTION_IDS:
        cdm = load_cdm(f"{INST_CDM_PATH}/{inst_sol_id}.cdm")
        print(f"{cdm.name}: {num_elements(cdm)}")


def num_elements(cdm: ClassDiagram):
    "Return the number of elements in the given class diagram according to the formula used in the experiment."
    result = 1  # 1 for Incomplete containment tree
    for e in cdm.eAllContents():
        match e:
            case (CDByte() | CDBoolean() | CDChar() | CDDouble() | CDFloat() | CDInt() | CDLong() | CDString() | CDAny()
                  | CDArray()):
                pass
            case AssociationEnd():
                result += 2  # role name, multiplicity
            case Classifier() as c:
                result += 1 + len(c.superTypes)
            case _:
                result += 1
    return result


if __name__ == "__main__":
    print_inst_sol_num_elements()
    produce_common_spreadsheet()
